import pandas as pd
import json
from pathlib import Path
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, Reference

BASE_DIR = Path(__file__).resolve().parent.parent
DATASET_PATH = BASE_DIR / "cleaned_dataset.csv"
SETTINGS_PATH = BASE_DIR / "settings.json"
OUTPUT_PATH = BASE_DIR / "report.xlsx"

def safe_eval_equation(df: pd.DataFrame, equation: str) -> pd.Series:
    allowed = {col: df[col] for col in df.columns if col != "date"}
    pattern = r"^[\dkpi+\-*/().\s]+$"
    if not re.match(pattern, equation):
        raise ValueError(f"Invalid equation: {equation}")
    return eval(equation, {"__builtins__": {}}, allowed)

def generate_excel_report():
    df = pd.read_csv(DATASET_PATH, parse_dates=["date"])

    with open(SETTINGS_PATH, "r") as f:
        settings = json.load(f)
        if not settings.get("line_chart"):
            settings["line_chart"] = {"type": "list", "value": []}
        if not settings.get("bar_chart"):
            settings["bar_chart"] = {"type": "list", "value": []}

    days_back = settings.get("days_back", 0)
    if days_back > 0:
        df = df.sort_values("date").tail(days_back)

    line_data = pd.DataFrame()
    bar_data = pd.DataFrame()
    
    if settings["line_chart"]["type"] == "equation":
        eq = settings["line_chart"]["value"]
        if eq:
            df["line_equation"] = safe_eval_equation(df, eq)
            line_data = df[["line_equation"]]
    elif settings["line_chart"]["type"] == "list" and settings["line_chart"]["value"]:
        line_data = df[settings["line_chart"]["value"]]

    if settings["bar_chart"]["type"] == "equation":
        eq = settings["bar_chart"]["value"]
        if eq:
            df["bar_equation"] = safe_eval_equation(df, eq)
            bar_data = df[["bar_equation"]]
    elif settings["bar_chart"]["type"] == "list" and settings["bar_chart"]["value"]:
        bar_data = df[settings["bar_chart"]["value"]]

    combined = pd.concat([df[["date"]], line_data, bar_data], axis=1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    for r in dataframe_to_rows(combined, index=False, header=True):
        ws.append(r)

    line_chart = LineChart()
    line_chart.style = 2
    if line_data.shape[1] > 0:
        line_chart.add_data(
            Reference(ws, min_col=2, max_col=1 + line_data.shape[1],
                      min_row=1, max_row=ws.max_row),
            titles_from_data=True,
        )
        line_chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
        line_chart.x_axis.number_format = "dd-mmm"
        line_chart.y_axis.number_format = "0"

    bar_chart = BarChart()
    bar_chart.style = 2
    if bar_data.shape[1] > 0:
        bar_chart.add_data(
            Reference(ws, min_col=2 + line_data.shape[1],
                      max_col=1 + line_data.shape[1] + bar_data.shape[1],
                      min_row=1, max_row=ws.max_row),
            titles_from_data=True,
        )
        bar_chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
        bar_chart.y_axis.axId = 200
        bar_chart.y_axis.number_format = "0"
        line_chart += bar_chart

    line_chart.legend.position = "r"
    ws.add_chart(line_chart, "H10")

    if OUTPUT_PATH.exists():
        OUTPUT_PATH.unlink()
    wb.save(OUTPUT_PATH)
    print(f"Excel report generated: {OUTPUT_PATH}")
